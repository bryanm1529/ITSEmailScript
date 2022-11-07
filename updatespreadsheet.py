import sys

# from colorama import Fore, init, Back, Style
from openpyxl import load_workbook
import re
import quickstart
import openpyxl
from string import digits
from openpyxl.styles import Alignment


"""Reading excel file using openpyxl"""
# load excel file
workbook = load_workbook(filename="\ITS\Projects\Banner\Upgrades_Releases")
workbook.active = workbook["Ellucian"]
# ws
sheet = workbook.active

# Colors declarations
my_red = openpyxl.styles.colors.Color(rgb="00FF0000")
red_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=my_red)

my_green = openpyxl.styles.colors.Color(rgb="47A34D")
green_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=my_green)

my_yellow = openpyxl.styles.colors.Color(rgb="00FFFF00")
yellow_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=my_yellow)


def update_sheet(banner_releases, ws):
    not_counter = 0
    repost_counter = 0
    new_counter = 0
    is_counter = 0
    for i, (release, link) in enumerate(banner_releases):
        cell_tracker = 0
        for cell in ws["A"]:
            if cell.value == None:
                break
            cell_tracker += 1
            repost_date_cell = ws["F" + str(cell_tracker)]
            link_cell = ws["I" + str(cell_tracker)]
            banner_releases_cell = cell
            exclusion_symbols = [
                "SC",
                "TCC",
                "UK",
                "ASCGEN",
                "Australia",
                "REPT",
                "CALB",
            ]
            # Checking  if its RCNJ related
            if any(word in str(release).split() for word in exclusion_symbols):

                not_rcnj_related(release, link, ws)
                not_counter += 1
                break

            print(banner_releases_cell.value)
            print(release)

            # Case where we only need to update repost date
            if str(banner_releases_cell.value) == str(release):
                repost_date_cell.value = str(quickstart.final_date)
                repost_date_cell.alignment = Alignment(horizontal="right")
                link_cell.value = link
                link_cell.alignment = Alignment(horizontal="left")
                repost_counter += 1
                break

            # Case where it could be rcnj related
            else:
                admin_answer = input(
                    f"Is the release  {release} RCNJ-related? Please Enter (y/n) or anything else if you are unsure "
                )
                if admin_answer.lower() == "n":
                    not_rcnj_related(release, link, ws)
                    not_counter += 1
                    break
                elif admin_answer.lower() == "y":
                    is_rcnj_related(release, link, ws)
                    is_counter += 1
                    break
                else:
                    unsure_rcnj_related(release, link, ws)
                    new_counter += 1
                    break

    print(
        f"Repost counter = {repost_counter}, not_counter = {not_counter}, new_counter = {new_counter}, is_counter = {is_counter}"
    )
    # # Case where we need to update the version number
    # elif str(banner_releases_cell.value.translate(digits)) == release.translate(digits):
    #     print("This is the version numbers: " ,str(banner_releases_cell.value.translate(digits)))
    #     print(release.translate(digits))
    #     # banner_releases_cell.value = release
    #     # repost_date_cell.value = (quickstart.final_date)
    #     # link_cell.value = link
    #     break


# function for getting last row in sheet so we can add a new release in the sheet
# gets the last empty row
def get_last_row(ws):
    last_row = ws.max_row

    while ws.cell(column=1, row=last_row).value is None and last_row > 0:
        last_row -= 1
    last_col_a_value = ws.cell(column=1, row=last_row).value
    return last_row + 1


def not_rcnj_related(release, link, ws):
    current_row = str(get_last_row(ws))
    insert_release(release, ws)
    for cell in ws[current_row + ":" + current_row]:
        cell.fill = red_fill

    rcnj_cell = ws["B" + current_row]
    rcnj_cell.value = "No"

    email_date_cell = ws["D" + current_row]
    email_date_cell.value = str(quickstart.final_date)
    email_date_cell.alignment = Alignment(horizontal="right")

    link_cell = ws["I" + current_row]
    link_cell.value = link


def is_rcnj_related(release, link, ws):
    current_row = str(get_last_row(ws))
    insert_release(release, ws)
    for cell in ws[current_row + ":" + current_row]:
        cell.fill = green_fill

    rcnj_cell = ws["B" + current_row]
    rcnj_cell.value = "Yes"

    email_date_cell = ws["D" + current_row]
    email_date_cell.value = str(quickstart.final_date)
    email_date_cell.alignment = Alignment(horizontal="right")

    link_cell = ws["I" + current_row]
    link_cell.value = link


def unsure_rcnj_related(release, link, ws):
    current_row = str(get_last_row(ws))
    insert_release(release, ws)
    for cell in ws[current_row + ":" + current_row]:
        cell.fill = yellow_fill

    rcnj_cell = ws["B" + current_row]
    rcnj_cell.value = "?"

    email_date_cell = ws["D" + current_row]
    email_date_cell.value = str(quickstart.final_date)
    email_date_cell.alignment = Alignment(horizontal="right")

    link_cell = ws["I" + current_row]
    link_cell.value = link


def insert_email_date(email_cell, current_row, ws):
    email_cell = ws["D" + current_row]
    email_cell.value = str(quickstart.final_date)
    email_cell.alignment = Alignment(horizontal="right")


def insert_release(release, ws):
    last_row = get_last_row(ws)
    new_cell = ws.cell(column=1, row=last_row)
    if new_cell.value is None:
        new_cell.value = release
    else:
        print("Cell is already populated")


release_test = [
    (
        "Human Resources 8.19.2",
        "https://ellucian.force.com/clients/s/releases/a111M00000RP3OY/ba-hr-8192",
    ),
    (
        "Finance 8.13.1.3",
        "https://ellucian.force.com/clients/s/releases/a111M00000RP3OY/ba-hr-8192",
    ),
    (
        "Testing TCC 8.13.1.3",
        "https://ellucian.force.com/clients/s/releases/a111M00000RP3OY/ba-hr-8192",
    ),
]
update_sheet(quickstart.final_banner_releases, sheet)


# modify the desired cell
# sheet["A1"] = "Full Name"


# Enumerate the cells in the  row we want to update to then color them if the banner is ramappo related
# for cell in ws["2:2"]:
#     cell.style = 'red_italic'

# trying to iterate over banner releases column values
# for row in sheet.iter_rows(min_row=1, max_col=1, max_row=167, values_only=True):
#     for release in banner_releases:
#         #Want to update sheet if release is in row identically
#         if release == row:

#         #want to regex for release without version number
#         # elif release == revised_release:

#         #Update sheet with new releases if they are not there at all
#         else:
#             #call function to get last row
# save the file
workbook.save(filename="\ITS\Projects\Banner\Upgrades_Releases.xlsx")
